"""
Módulo de Banco de Dados - FAI-SQL Fluent
Funções utilitárias para conexão e execução de queries SQL Server via pyodbc.
"""

import csv
import os
from datetime import datetime

import pyodbc


def testar_conexao(ip: str, usuario: str, senha: str, banco: str) -> tuple[bool, str]:
    """Testa se uma conexão SQL Server é válida."""
    string_conexao = (
        f"DRIVER={{SQL Server}};"
        f"SERVER={ip};"
        f"DATABASE={banco};"
        f"UID={usuario};"
        f"PWD={senha};"
    )
    try:
        conexao = pyodbc.connect(string_conexao, timeout=5)
        conexao.close()
        return True, "Conexão OK!"
    except Exception as e:
        return False, str(e)


def executar_query(dados_conexao: dict, comando: str) -> dict:
    """
    Executa um comando SQL e retorna o resultado.
    
    Retorna dict com:
        - sucesso (bool)
        - tipo: "select" | "dml"
        - colunas: list[str] (somente para select)
        - dados: list[list[str]] (somente para select)
        - registros: int
        - mensagem: str
    """
    string_conexao = (
        f"DRIVER={{SQL Server}};"
        f"SERVER={dados_conexao.get('ip', '')};"
        f"DATABASE={dados_conexao.get('banco', '')};"
        f"UID={dados_conexao.get('usuario', '')};"
        f"PWD={dados_conexao.get('senha', '')};"
    )

    try:
        conexao = pyodbc.connect(string_conexao)
        cursor = conexao.cursor()
        cursor.execute(comando)

        if comando.strip().upper().startswith("SELECT"):
            resultados = cursor.fetchall()
            colunas = [col[0] for col in cursor.description]
            dados = []
            for linha in resultados:
                dados.append([str(v) if v is not None else "NULL" for v in linha])

            cursor.close()
            conexao.close()

            return {
                "sucesso": True,
                "tipo": "select",
                "colunas": colunas,
                "dados": dados,
                "registros": len(resultados),
                "mensagem": f"{len(resultados)} registro(s) retornado(s)",
            }
        else:
            conexao.commit()
            cursor.close()
            conexao.close()
            return {
                "sucesso": True,
                "tipo": "dml",
                "colunas": [],
                "dados": [],
                "registros": 0,
                "mensagem": "Comando executado com sucesso!",
            }

    except Exception as e:
        return {
            "sucesso": False,
            "tipo": "error",
            "colunas": [],
            "dados": [],
            "registros": 0,
            "mensagem": str(e),
        }


def exportar_csv(caminho: str, colunas: list[str], dados: list[list[str]]):
    """Exporta dados para CSV com separador ;"""
    with open(caminho, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerow(colunas)
        for linha in dados:
            writer.writerow(linha)


def exportar_excel(caminho: str, colunas: list[str], dados: list[list[str]]):
    """Exporta dados para Excel com formatação profissional."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Dados"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Cabeçalho
    for col_idx, coluna in enumerate(colunas, 1):
        cell = ws.cell(row=1, column=col_idx, value=coluna)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Dados
    for row_idx, linha in enumerate(dados, 2):
        for col_idx, valor in enumerate(linha, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    # Ajustar largura das colunas
    for col_idx, coluna in enumerate(colunas, 1):
        max_length = len(str(coluna))
        for linha in dados:
            if col_idx <= len(linha):
                max_length = max(max_length, len(str(linha[col_idx - 1])))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(
            max_length + 2, 50
        )

    wb.save(caminho)
