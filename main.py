import tkinter as tk
from tkinter import ttk, messagebox, font, filedialog
import pyodbc
import json
import os
import csv
import re
from datetime import datetime

# ============================================
# CONFIGURAÇÃO DE TEMAS
# ============================================

TEMAS = {
    "claro": {
        "bg": "#f5f5f5",
        "fg": "#1a1a1a",
        "bg_secondary": "#ffffff",
        "accent": "#0066cc",
        "accent_hover": "#0052a3",
        "border": "#d0d0d0",
        "success": "#28a745",
        "error": "#dc3545",
        "treeview_bg": "#ffffff",
        "treeview_fg": "#1a1a1a",
        "treeview_selected": "#0066cc",
        "entry_bg": "#ffffff",
        "button_bg": "#0066cc",
        "button_fg": "#ffffff",
        "menu_bg": "#f5f5f5",
        "menu_fg": "#1a1a1a",
        # Syntax Highlighting
        "sql_keyword": "#0000ff",
        "sql_function": "#9932cc",
        "sql_string": "#008000",
        "sql_number": "#ff6600",
        "sql_comment": "#808080",
        "sql_operator": "#cc0066"
    },
    "escuro": {
        "bg": "#1e1e1e",
        "fg": "#e0e0e0",
        "bg_secondary": "#2d2d2d",
        "accent": "#4da6ff",
        "accent_hover": "#3d8bd9",
        "border": "#404040",
        "success": "#4caf50",
        "error": "#f44336",
        "treeview_bg": "#2d2d2d",
        "treeview_fg": "#e0e0e0",
        "treeview_selected": "#4da6ff",
        "entry_bg": "#3d3d3d",
        "button_bg": "#4da6ff",
        "button_fg": "#1e1e1e",
        "menu_bg": "#2d2d2d",
        "menu_fg": "#e0e0e0",
        # Syntax Highlighting
        "sql_keyword": "#569cd6",
        "sql_function": "#dcdcaa",
        "sql_string": "#ce9178",
        "sql_number": "#b5cea8",
        "sql_comment": "#6a9955",
        "sql_operator": "#d4d4d4"
    }
}

tema_atual = "claro"

def carregar_preferencia_tema():
    """Carrega a preferência de tema do arquivo de configuração"""
    global tema_atual
    if os.path.exists("config.json"):
        try:
            with open("config.json", "r") as arquivo:
                dados = json.load(arquivo)
                tema_atual = dados.get("tema", "claro")
        except:
            tema_atual = "claro"

def salvar_preferencia_tema():
    """Salva a preferência de tema no arquivo de configuração"""
    dados = {}
    if os.path.exists("config.json"):
        try:
            with open("config.json", "r") as arquivo:
                dados = json.load(arquivo)
        except:
            pass
    dados["tema"] = tema_atual
    with open("config.json", "w") as arquivo:
        json.dump(dados, arquivo, indent=4)

def aplicar_tema():
    """Aplica o tema atual a todos os widgets"""
    tema = TEMAS[tema_atual]
    
    # Configurar estilo ttk
    style = ttk.Style()
    style.theme_use('clam')
    
    # Frame
    style.configure("TFrame", background=tema["bg"])
    
    # Label
    style.configure("TLabel", 
                    background=tema["bg"], 
                    foreground=tema["fg"],
                    font=("Segoe UI", 10))
    
    style.configure("Title.TLabel",
                    background=tema["bg"],
                    foreground=tema["fg"],
                    font=("Segoe UI", 12, "bold"))
    
    # Button
    style.configure("TButton",
                    background=tema["button_bg"],
                    foreground=tema["button_fg"],
                    font=("Segoe UI", 10),
                    padding=(15, 8))
    
    style.configure("Accent.TButton",
                    background=tema["accent"],
                    foreground=tema["button_fg"],
                    font=("Segoe UI", 10, "bold"),
                    padding=(20, 10))
    
    style.map("TButton",
              background=[("active", tema["accent_hover"]), ("pressed", tema["accent_hover"])])
    
    style.map("Accent.TButton",
              background=[("active", tema["accent_hover"]), ("pressed", tema["accent_hover"])])
    
    # Combobox
    style.configure("TCombobox",
                    fieldbackground=tema["entry_bg"],
                    background=tema["entry_bg"],
                    foreground=tema["fg"],
                    arrowcolor=tema["fg"],
                    padding=5)
    
    style.map("TCombobox",
              fieldbackground=[("readonly", tema["entry_bg"])],
              selectbackground=[("readonly", tema["accent"])])
    
    # Treeview
    style.configure("Treeview",
                    background=tema["treeview_bg"],
                    foreground=tema["treeview_fg"],
                    fieldbackground=tema["treeview_bg"],
                    rowheight=28,
                    font=("Segoe UI", 9))
    
    style.configure("Treeview.Heading",
                    background=tema["bg_secondary"],
                    foreground=tema["fg"],
                    font=("Segoe UI", 10, "bold"),
                    padding=5)
    
    style.map("Treeview",
              background=[("selected", tema["treeview_selected"])],
              foreground=[("selected", "#ffffff")])
    
    # Entry
    style.configure("TEntry",
                    fieldbackground=tema["entry_bg"],
                    foreground=tema["fg"],
                    padding=8)
    
    # Scrollbar
    style.configure("Vertical.TScrollbar",
                    background=tema["bg_secondary"],
                    troughcolor=tema["bg"],
                    arrowcolor=tema["fg"])
    
    style.configure("Horizontal.TScrollbar",
                    background=tema["bg_secondary"],
                    troughcolor=tema["bg"],
                    arrowcolor=tema["fg"])
    
    # LabelFrame
    style.configure("TLabelframe",
                    background=tema["bg"],
                    foreground=tema["fg"])
    
    style.configure("TLabelframe.Label",
                    background=tema["bg"],
                    foreground=tema["fg"],
                    font=("Segoe UI", 10, "bold"))
    
    # Aplicar tema à janela principal
    root.configure(bg=tema["bg"])
    
    # Atualizar widgets tk (não ttk)
    atualizar_widgets_tk(root, tema)

def atualizar_widgets_tk(widget, tema):
    """Atualiza widgets tk padrão com o tema"""
    try:
        widget_type = widget.winfo_class()
        
        if widget_type == "Text":
            widget.configure(
                bg=tema["entry_bg"],
                fg=tema["fg"],
                insertbackground=tema["fg"],
                selectbackground=tema["accent"],
                selectforeground="#ffffff",
                relief="flat",
                borderwidth=2,
                highlightthickness=1,
                highlightbackground=tema["border"],
                highlightcolor=tema["accent"]
            )
        elif widget_type == "Menu":
            widget.configure(
                bg=tema["menu_bg"],
                fg=tema["menu_fg"],
                activebackground=tema["accent"],
                activeforeground="#ffffff",
                borderwidth=0
            )
    except tk.TclError:
        pass
    
    # Recursivamente atualizar filhos
    for child in widget.winfo_children():
        atualizar_widgets_tk(child, tema)

def alternar_tema():
    """Alterna entre tema claro e escuro"""
    global tema_atual
    tema_atual = "escuro" if tema_atual == "claro" else "claro"
    aplicar_tema()
    salvar_preferencia_tema()
    atualizar_texto_menu_tema()

def atualizar_texto_menu_tema():
    """Atualiza o texto do menu de tema"""
    novo_texto = "🌙 Tema Escuro" if tema_atual == "claro" else "☀️ Tema Claro"
    menu_sistema.entryconfig(4, label=novo_texto)

def sair():
    root.quit()

# ============================================
# GERENCIAMENTO DE CONEXÕES
# ============================================

conexao_ativa = None  # Nome da conexão ativa

def carregar_conexoes():
    """Carrega todas as conexões do config.json"""
    if os.path.exists("config.json"):
        try:
            with open("config.json", "r") as f:
                dados = json.load(f)
                # Migrar formato antigo para novo se necessário
                if "conexoes" not in dados:
                    if "ip" in dados:
                        # Migrar conexão existente
                        dados["conexoes"] = [{
                            "nome": "Padrão",
                            "ip": dados.get("ip", ""),
                            "usuario": dados.get("usuario", ""),
                            "senha": dados.get("senha", ""),
                            "banco": dados.get("banco", "")
                        }]
                        dados["conexao_ativa"] = "Padrão"
                        # Manter tema
                        tema = dados.get("tema", "claro")
                        dados["tema"] = tema
                        # Salvar migração
                        with open("config.json", "w") as fw:
                            json.dump(dados, fw, indent=4)
                return dados.get("conexoes", [])
        except:
            return []
    return []

def obter_conexao_ativa():
    """Retorna os dados da conexão ativa"""
    if os.path.exists("config.json"):
        try:
            with open("config.json", "r") as f:
                dados = json.load(f)
                nome_ativa = dados.get("conexao_ativa", "")
                conexoes = dados.get("conexoes", [])
                for conn in conexoes:
                    if conn.get("nome") == nome_ativa:
                        return conn
                # Fallback para formato antigo
                if "ip" in dados and "conexoes" not in dados:
                    return dados
        except:
            pass
    return None

def definir_conexao_ativa(nome):
    """Define qual conexão é a ativa"""
    global conexao_ativa
    conexao_ativa = nome
    if os.path.exists("config.json"):
        try:
            with open("config.json", "r") as f:
                dados = json.load(f)
            dados["conexao_ativa"] = nome
            with open("config.json", "w") as f:
                json.dump(dados, f, indent=4)
            atualizar_label_conexao()
        except:
            pass

def atualizar_label_conexao():
    """Atualiza o label que mostra a conexão ativa"""
    try:
        conn = obter_conexao_ativa()
        if conn:
            texto = f"🔗 {conn.get('nome', 'Sem nome')} ({conn.get('banco', '')})"
        else:
            texto = "⚠️ Nenhuma conexão configurada"
        label_conexao_ativa.config(text=texto)
    except:
        pass

def abrir_tela_conexao():
    """Abre a tela de gerenciamento de conexões"""
    tema = TEMAS[tema_atual]
    
    tela_conexao = tk.Toplevel(root)
    tela_conexao.title("⚙️ Gerenciador de Conexões")
    tela_conexao.geometry("600x650")
    tela_conexao.configure(bg=tema["bg"])
    tela_conexao.resizable(False, False)
    
    tela_conexao.transient(root)
    tela_conexao.grab_set()
    
    main_frame = ttk.Frame(tela_conexao, padding=20)
    main_frame.pack(fill="both", expand=True)
    
    titulo = ttk.Label(main_frame, text="Gerenciador de Conexões", style="Title.TLabel")
    titulo.pack(pady=(0, 15))
    
    # Frame para lista de conexões
    frame_lista = ttk.LabelFrame(main_frame, text="Conexões Salvas", padding=10)
    frame_lista.pack(fill="x", pady=(0, 15))
    
    lista_conexoes = tk.Listbox(frame_lista, height=5, font=("Segoe UI", 10),
                                 bg=tema["entry_bg"], fg=tema["fg"],
                                 selectbackground=tema["accent"],
                                 selectforeground="#ffffff",
                                 relief="flat", highlightthickness=1,
                                 highlightbackground=tema["border"])
    lista_conexoes.pack(fill="x", pady=(0, 10))
    
    def atualizar_lista():
        lista_conexoes.delete(0, tk.END)
        conexoes = carregar_conexoes()
        for conn in conexoes:
            lista_conexoes.insert(tk.END, f"{conn.get('nome', 'Sem nome')} - {conn.get('banco', '')}")
    
    def ao_selecionar_conexao(event=None):
        sel = lista_conexoes.curselection()
        if sel:
            conexoes = carregar_conexoes()
            if sel[0] < len(conexoes):
                conn = conexoes[sel[0]]
                entry_nome.delete(0, tk.END)
                entry_nome.insert(0, conn.get("nome", ""))
                entry_ip.delete(0, tk.END)
                entry_ip.insert(0, conn.get("ip", ""))
                entry_usuario.delete(0, tk.END)
                entry_usuario.insert(0, conn.get("usuario", ""))
                entry_senha.delete(0, tk.END)
                entry_senha.insert(0, conn.get("senha", ""))
                entry_banco.delete(0, tk.END)
                entry_banco.insert(0, conn.get("banco", ""))
    
    lista_conexoes.bind("<<ListboxSelect>>", ao_selecionar_conexao)
    
    # Botões de ação da lista
    frame_btns_lista = ttk.Frame(frame_lista)
    frame_btns_lista.pack(fill="x")
    
    def usar_conexao():
        sel = lista_conexoes.curselection()
        if not sel:
            messagebox.showwarning("⚠️ Atenção", "Selecione uma conexão!", parent=tela_conexao)
            return
        conexoes = carregar_conexoes()
        if sel[0] < len(conexoes):
            conn = conexoes[sel[0]]
            definir_conexao_ativa(conn.get("nome"))
            messagebox.showinfo("✅ Sucesso", f"Conexão '{conn.get('nome')}' ativada!", parent=tela_conexao)
    
    def excluir_conexao():
        sel = lista_conexoes.curselection()
        if not sel:
            messagebox.showwarning("⚠️ Atenção", "Selecione uma conexão!", parent=tela_conexao)
            return
        if not messagebox.askyesno("🗑️ Confirmar", "Excluir esta conexão?", parent=tela_conexao):
            return
        try:
            with open("config.json", "r") as f:
                dados = json.load(f)
            conexoes = dados.get("conexoes", [])
            if sel[0] < len(conexoes):
                del conexoes[sel[0]]
                dados["conexoes"] = conexoes
                with open("config.json", "w") as f:
                    json.dump(dados, f, indent=4)
                atualizar_lista()
                messagebox.showinfo("✅ Sucesso", "Conexão excluída!", parent=tela_conexao)
        except Exception as e:
            messagebox.showerror("❌ Erro", str(e), parent=tela_conexao)
    
    ttk.Button(frame_btns_lista, text="✅ Usar", command=usar_conexao).pack(side="left", padx=(0, 5))
    ttk.Button(frame_btns_lista, text="🗑️ Excluir", command=excluir_conexao).pack(side="left")
    
    # Frame para edição/adição
    frame_form = ttk.LabelFrame(main_frame, text="Dados da Conexão", padding=10)
    frame_form.pack(fill="x", pady=(0, 15))
    
    # Campos
    campos = [
        ("Nome:", "entry_nome"),
        ("Servidor/IP:", "entry_ip"),
        ("Usuário:", "entry_usuario"),
        ("Senha:", "entry_senha"),
        ("Banco de Dados:", "entry_banco")
    ]
    
    entries = {}
    for label_text, entry_name in campos:
        frame = ttk.Frame(frame_form)
        frame.pack(fill="x", pady=4)
        
        label = ttk.Label(frame, text=label_text, width=15)
        label.pack(side="left")
        
        show_char = "*" if "senha" in entry_name.lower() else ""
        entry = tk.Entry(frame, font=("Segoe UI", 10), show=show_char,
                        bg=tema["entry_bg"], fg=tema["fg"],
                        insertbackground=tema["fg"], relief="flat",
                        highlightthickness=1, highlightbackground=tema["border"],
                        highlightcolor=tema["accent"])
        entry.pack(side="left", fill="x", expand=True, ipady=4)
        entries[entry_name] = entry
    
    entry_nome = entries["entry_nome"]
    entry_ip = entries["entry_ip"]
    entry_usuario = entries["entry_usuario"]
    entry_senha = entries["entry_senha"]
    entry_banco = entries["entry_banco"]
    
    # Botões de ação
    frame_acoes = ttk.Frame(main_frame)
    frame_acoes.pack(fill="x")
    
    def testar_conexao():
        string_conexao = (
            f"DRIVER={{SQL Server}};"
            f"SERVER={entry_ip.get()};"
            f"DATABASE={entry_banco.get()};"
            f"UID={entry_usuario.get()};"
            f"PWD={entry_senha.get()};"
        )
        try:
            conexao = pyodbc.connect(string_conexao, timeout=5)
            conexao.close()
            messagebox.showinfo("✅ Sucesso", "Conexão OK!", parent=tela_conexao)
        except Exception as e:
            messagebox.showerror("❌ Erro", f"Falha:\n{str(e)}", parent=tela_conexao)
    
    def salvar_conexao():
        nome = entry_nome.get().strip()
        if not nome:
            messagebox.showwarning("⚠️ Atenção", "Informe um nome para a conexão!", parent=tela_conexao)
            return
        
        nova_conn = {
            "nome": nome,
            "ip": entry_ip.get(),
            "usuario": entry_usuario.get(),
            "senha": entry_senha.get(),
            "banco": entry_banco.get()
        }
        
        try:
            dados = {}
            if os.path.exists("config.json"):
                with open("config.json", "r") as f:
                    dados = json.load(f)
            
            conexoes = dados.get("conexoes", [])
            
            # Verificar se já existe e atualizar
            encontrado = False
            for i, conn in enumerate(conexoes):
                if conn.get("nome") == nome:
                    conexoes[i] = nova_conn
                    encontrado = True
                    break
            
            if not encontrado:
                conexoes.append(nova_conn)
            
            dados["conexoes"] = conexoes
            if not dados.get("conexao_ativa"):
                dados["conexao_ativa"] = nome
            
            with open("config.json", "w") as f:
                json.dump(dados, f, indent=4)
            
            atualizar_lista()
            atualizar_label_conexao()
            messagebox.showinfo("✅ Sucesso", "Conexão salva!", parent=tela_conexao)
            
        except Exception as e:
            messagebox.showerror("❌ Erro", str(e), parent=tela_conexao)
    
    def limpar_campos():
        for entry in entries.values():
            entry.delete(0, tk.END)
    
    ttk.Button(frame_acoes, text="🔌 Testar", command=testar_conexao).pack(side="left", padx=(0, 5))
    ttk.Button(frame_acoes, text="💾 Salvar", style="Accent.TButton", command=salvar_conexao).pack(side="left", padx=(0, 5))
    ttk.Button(frame_acoes, text="🆕 Novo", command=limpar_campos).pack(side="left")
    
    atualizar_lista()

def abrir_tela_adicionar_comando():
    tema = TEMAS[tema_atual]
    
    tela_adicionar_comando = tk.Toplevel(root)
    tela_adicionar_comando.title("➕ Adicionar Comando")
    tela_adicionar_comando.geometry("700x500")
    tela_adicionar_comando.configure(bg=tema["bg"])
    tela_adicionar_comando.minsize(500, 400)
    
    tela_adicionar_comando.transient(root)
    tela_adicionar_comando.grab_set()
    
    main_frame = ttk.Frame(tela_adicionar_comando, padding=25)
    main_frame.pack(fill="both", expand=True)
    
    titulo = ttk.Label(main_frame, text="Adicionar Novo Comando", style="Title.TLabel")
    titulo.pack(pady=(0, 20))

    def salvar_comando():
        nome_comando = entry_nome_comando.get()
        comando = caixa_texto_comando.get("1.0", tk.END).strip()

        if not nome_comando or not comando:
            messagebox.showwarning("⚠️ Atenção", "Preencha todos os campos!", parent=tela_adicionar_comando)
            return

        novo_comando = {
            "nome": nome_comando,
            "comando": comando
        }

        if os.path.exists("comandos.json"):
            with open("comandos.json", "r") as arquivo:
                comandos = json.load(arquivo)
        else:
            comandos = []

        comandos.append(novo_comando)

        with open("comandos.json", "w") as arquivo:
            json.dump(comandos, arquivo, indent=4)

        messagebox.showinfo("✅ Sucesso", "Comando salvo!", parent=tela_adicionar_comando)
        tela_adicionar_comando.destroy()
        carregar_comandos_no_combobox()

    # Nome do comando
    frame_nome = ttk.Frame(main_frame)
    frame_nome.pack(fill="x", pady=(0, 15))
    
    label_nome = ttk.Label(frame_nome, text="Nome do Comando:")
    label_nome.pack(anchor="w")
    
    entry_nome_comando = tk.Entry(frame_nome, font=("Segoe UI", 10),
                                  bg=tema["entry_bg"], fg=tema["fg"],
                                  insertbackground=tema["fg"], relief="flat",
                                  highlightthickness=1, highlightbackground=tema["border"],
                                  highlightcolor=tema["accent"])
    entry_nome_comando.pack(fill="x", pady=(4, 0), ipady=6)
    
    # Comando SQL
    frame_comando = ttk.Frame(main_frame)
    frame_comando.pack(fill="both", expand=True, pady=(0, 15))
    
    label_comando = ttk.Label(frame_comando, text="Comando SQL:")
    label_comando.pack(anchor="w")
    
    caixa_texto_comando = tk.Text(frame_comando, font=("Consolas", 11), wrap=tk.WORD,
                                   bg=tema["entry_bg"], fg=tema["fg"],
                                   insertbackground=tema["fg"], relief="flat",
                                   highlightthickness=1, highlightbackground=tema["border"],
                                   highlightcolor=tema["accent"])
    caixa_texto_comando.pack(fill="both", expand=True, pady=(4, 0))

    btn_salvar = ttk.Button(main_frame, text="💾 Salvar Comando", 
                            style="Accent.TButton", command=salvar_comando)
    btn_salvar.pack(pady=(10, 0))

def abrir_tela_editar_comando():
    """Abre janela para editar o comando selecionado"""
    nome_comando_selecionado = combo_comandos.get()
    
    if not nome_comando_selecionado:
        messagebox.showwarning("⚠️ Atenção", "Selecione um comando para editar!")
        return
    
    tema = TEMAS[tema_atual]
    
    tela_editar = tk.Toplevel(root)
    tela_editar.title("✏️ Editar Comando")
    tela_editar.geometry("700x550")
    tela_editar.configure(bg=tema["bg"])
    tela_editar.minsize(500, 500)
    
    tela_editar.transient(root)
    tela_editar.grab_set()
    
    main_frame = ttk.Frame(tela_editar, padding=25)
    main_frame.pack(fill="both", expand=True)
    
    titulo = ttk.Label(main_frame, text="Editar Comando", style="Title.TLabel")
    titulo.pack(pady=(0, 20))
    
    # Carregar dados do comando
    comando_atual = None
    indice_comando = -1
    if os.path.exists("comandos.json"):
        with open("comandos.json", "r") as arquivo:
            comandos = json.load(arquivo)
            for i, cmd in enumerate(comandos):
                if cmd["nome"] == nome_comando_selecionado:
                    comando_atual = cmd
                    indice_comando = i
                    break
    
    if comando_atual is None:
        messagebox.showerror("❌ Erro", "Comando não encontrado!", parent=tela_editar)
        tela_editar.destroy()
        return
    
    def salvar_edicao():
        novo_nome = entry_nome_comando.get()
        novo_comando = caixa_texto_edit.get("1.0", tk.END).strip()
        
        if not novo_nome or not novo_comando:
            messagebox.showwarning("⚠️ Atenção", "Preencha todos os campos!", parent=tela_editar)
            return
        
        # Atualizar no arquivo
        with open("comandos.json", "r") as arquivo:
            comandos = json.load(arquivo)
        
        comandos[indice_comando] = {
            "nome": novo_nome,
            "comando": novo_comando
        }
        
        with open("comandos.json", "w") as arquivo:
            json.dump(comandos, arquivo, indent=4)
        
        messagebox.showinfo("✅ Sucesso", "Comando atualizado!", parent=tela_editar)
        tela_editar.destroy()
        carregar_comandos_no_combobox()
        
        # Atualizar seleção se o nome mudou
        if novo_nome != nome_comando_selecionado:
            combo_comandos.set(novo_nome)
        ao_selecionar_comando(None)
    
    # Nome do comando
    frame_nome = ttk.Frame(main_frame)
    frame_nome.pack(fill="x", pady=(0, 15))
    
    label_nome = ttk.Label(frame_nome, text="Nome do Comando:")
    label_nome.pack(anchor="w")
    
    entry_nome_comando = tk.Entry(frame_nome, font=("Segoe UI", 10),
                                  bg=tema["entry_bg"], fg=tema["fg"],
                                  insertbackground=tema["fg"], relief="flat",
                                  highlightthickness=1, highlightbackground=tema["border"],
                                  highlightcolor=tema["accent"])
    entry_nome_comando.pack(fill="x", pady=(4, 0), ipady=6)
    entry_nome_comando.insert(0, comando_atual["nome"])
    
    # Comando SQL
    frame_comando = ttk.Frame(main_frame)
    frame_comando.pack(fill="both", expand=True, pady=(0, 15))
    
    label_comando = ttk.Label(frame_comando, text="Comando SQL:")
    label_comando.pack(anchor="w")
    
    caixa_texto_edit = tk.Text(frame_comando, font=("Consolas", 11), wrap=tk.WORD,
                                bg=tema["entry_bg"], fg=tema["fg"],
                                insertbackground=tema["fg"], relief="flat",
                                highlightthickness=1, highlightbackground=tema["border"],
                                highlightcolor=tema["accent"])
    caixa_texto_edit.pack(fill="both", expand=True, pady=(4, 0))
    caixa_texto_edit.insert("1.0", comando_atual["comando"])
    
    btn_salvar = ttk.Button(main_frame, text="💾 Salvar Alterações", 
                            style="Accent.TButton", command=salvar_edicao)
    btn_salvar.pack(pady=(10, 0))

def excluir_comando():
    """Exclui o comando selecionado com confirmação"""
    nome_comando_selecionado = combo_comandos.get()
    
    if not nome_comando_selecionado:
        messagebox.showwarning("⚠️ Atenção", "Selecione um comando para excluir!")
        return
    
    # Confirmação
    confirmar = messagebox.askyesno(
        "🗑️ Confirmar Exclusão",
        f"Tem certeza que deseja excluir o comando:\n\n\"{nome_comando_selecionado}\"?\n\nEsta ação não pode ser desfeita."
    )
    
    if not confirmar:
        return
    
    # Remover do arquivo
    if os.path.exists("comandos.json"):
        with open("comandos.json", "r") as arquivo:
            comandos = json.load(arquivo)
        
        comandos = [cmd for cmd in comandos if cmd["nome"] != nome_comando_selecionado]
        
        with open("comandos.json", "w") as arquivo:
            json.dump(comandos, arquivo, indent=4)
        
        messagebox.showinfo("✅ Sucesso", "Comando excluído!")
        carregar_comandos_no_combobox()
        combo_comandos.set("")
        caixa_texto_comando.delete("1.0", tk.END)

def carregar_comandos_no_combobox():
    if os.path.exists("comandos.json"):
        with open("comandos.json", "r") as arquivo:
            comandos = json.load(arquivo)
            nomes_comandos = [comando["nome"] for comando in comandos]
            combo_comandos["values"] = nomes_comandos
    else:
        combo_comandos["values"] = []

def ao_selecionar_comando(event):
    nome_comando_selecionado = combo_comandos.get()

    if os.path.exists("comandos.json"):
        with open("comandos.json", "r") as arquivo:
            comandos = json.load(arquivo)
            for comando in comandos:
                if comando["nome"] == nome_comando_selecionado:
                    caixa_texto_comando.delete("1.0", tk.END)
                    caixa_texto_comando.insert("1.0", comando["comando"])
                    aplicar_syntax_highlighting()
                    break

# ============================================
# SYNTAX HIGHLIGHTING SQL
# ============================================

SQL_KEYWORDS = [
    "SELECT", "FROM", "WHERE", "AND", "OR", "NOT", "IN", "LIKE", "BETWEEN",
    "INSERT", "INTO", "VALUES", "UPDATE", "SET", "DELETE", "CREATE", "DROP",
    "ALTER", "TABLE", "INDEX", "VIEW", "DATABASE", "SCHEMA", "CONSTRAINT",
    "PRIMARY", "KEY", "FOREIGN", "REFERENCES", "UNIQUE", "CHECK", "DEFAULT",
    "NULL", "INNER", "LEFT", "RIGHT", "FULL", "OUTER", "JOIN", "ON", "AS",
    "ORDER", "BY", "ASC", "DESC", "GROUP", "HAVING", "DISTINCT", "TOP",
    "LIMIT", "OFFSET", "UNION", "ALL", "EXCEPT", "INTERSECT", "EXISTS",
    "CASE", "WHEN", "THEN", "ELSE", "END", "IF", "BEGIN", "COMMIT", "ROLLBACK",
    "TRANSACTION", "DECLARE", "EXEC", "EXECUTE", "PROCEDURE", "FUNCTION",
    "TRIGGER", "CURSOR", "FETCH", "OPEN", "CLOSE", "DEALLOCATE", "WITH",
    "CTE", "RECURSIVE", "PARTITION", "OVER", "ROW_NUMBER", "RANK", "DENSE_RANK",
    "IS", "TRUNCATE", "VARCHAR", "INT", "INTEGER", "FLOAT", "DECIMAL", "DATE",
    "DATETIME", "BIT", "TEXT", "NVARCHAR", "CHAR", "BIGINT", "SMALLINT"
]

SQL_FUNCTIONS = [
    "COUNT", "SUM", "AVG", "MIN", "MAX", "ABS", "CEILING", "FLOOR", "ROUND",
    "POWER", "SQRT", "LEN", "LENGTH", "SUBSTRING", "SUBSTR", "LEFT", "RIGHT",
    "LTRIM", "RTRIM", "TRIM", "UPPER", "LOWER", "REPLACE", "CHARINDEX",
    "CONCAT", "COALESCE", "ISNULL", "NULLIF", "CAST", "CONVERT", "GETDATE",
    "DATEADD", "DATEDIFF", "DATEPART", "YEAR", "MONTH", "DAY", "HOUR", "MINUTE",
    "SECOND", "NOW", "CURRENT_DATE", "CURRENT_TIME", "CURRENT_TIMESTAMP"
]

def configurar_tags_syntax(widget_texto):
    """Configura as tags de syntax highlighting no widget de texto"""
    tema = TEMAS[tema_atual]
    
    widget_texto.tag_configure("keyword", foreground=tema["sql_keyword"], font=("Consolas", 11, "bold"))
    widget_texto.tag_configure("function", foreground=tema["sql_function"])
    widget_texto.tag_configure("string", foreground=tema["sql_string"])
    widget_texto.tag_configure("number", foreground=tema["sql_number"])
    widget_texto.tag_configure("comment", foreground=tema["sql_comment"], font=("Consolas", 11, "italic"))
    widget_texto.tag_configure("operator", foreground=tema["sql_operator"])

def aplicar_syntax_highlighting(event=None):
    """Aplica syntax highlighting ao texto SQL"""
    try:
        # Remover todas as tags existentes
        for tag in ["keyword", "function", "string", "number", "comment", "operator"]:
            caixa_texto_comando.tag_remove(tag, "1.0", tk.END)
        
        texto = caixa_texto_comando.get("1.0", tk.END)
        
        # Destacar strings (entre aspas simples)
        for match in re.finditer(r"'[^']*'", texto):
            start = f"1.0+{match.start()}c"
            end = f"1.0+{match.end()}c"
            caixa_texto_comando.tag_add("string", start, end)
        
        # Destacar strings (entre aspas duplas)
        for match in re.finditer(r'"[^"]*"', texto):
            start = f"1.0+{match.start()}c"
            end = f"1.0+{match.end()}c"
            caixa_texto_comando.tag_add("string", start, end)
        
        # Destacar comentários de linha (-- ...)
        for match in re.finditer(r'--[^\n]*', texto):
            start = f"1.0+{match.start()}c"
            end = f"1.0+{match.end()}c"
            caixa_texto_comando.tag_add("comment", start, end)
        
        # Destacar comentários de bloco (/* ... */)
        for match in re.finditer(r'/\*[\s\S]*?\*/', texto):
            start = f"1.0+{match.start()}c"
            end = f"1.0+{match.end()}c"
            caixa_texto_comando.tag_add("comment", start, end)
        
        # Destacar números
        for match in re.finditer(r'\b\d+\.?\d*\b', texto):
            start = f"1.0+{match.start()}c"
            end = f"1.0+{match.end()}c"
            caixa_texto_comando.tag_add("number", start, end)
        
        # Destacar keywords SQL (case insensitive)
        for keyword in SQL_KEYWORDS:
            pattern = r'\b' + keyword + r'\b'
            for match in re.finditer(pattern, texto, re.IGNORECASE):
                start = f"1.0+{match.start()}c"
                end = f"1.0+{match.end()}c"
                caixa_texto_comando.tag_add("keyword", start, end)
        
        # Destacar funções SQL
        for func in SQL_FUNCTIONS:
            pattern = r'\b' + func + r'\s*\('
            for match in re.finditer(pattern, texto, re.IGNORECASE):
                start = f"1.0+{match.start()}c"
                end = f"1.0+{match.end() - 1}c"  # Não incluir o parêntese
                caixa_texto_comando.tag_add("function", start, end)
        
        # Destacar operadores
        for match in re.finditer(r'[=<>!]+|[+\-*/]', texto):
            start = f"1.0+{match.start()}c"
            end = f"1.0+{match.end()}c"
            caixa_texto_comando.tag_add("operator", start, end)
            
    except Exception:
        pass  # Ignorar erros de highlighting

# ============================================
# HISTÓRICO DE COMANDOS
# ============================================

HISTORICO_FILE = "historico.json"
MAX_HISTORICO = 100

def salvar_historico(comando, sucesso, mensagem="", registros=0):
    """Salva uma execução no histórico"""
    historico = []
    
    if os.path.exists(HISTORICO_FILE):
        try:
            with open(HISTORICO_FILE, "r", encoding="utf-8") as f:
                historico = json.load(f)
        except:
            historico = []
    
    nova_entrada = {
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "comando": comando[:500],  # Limitar tamanho
        "sucesso": sucesso,
        "mensagem": mensagem,
        "registros": registros
    }
    
    # Adicionar no início (mais recente primeiro)
    historico.insert(0, nova_entrada)
    
    # Limitar tamanho do histórico
    historico = historico[:MAX_HISTORICO]
    
    with open(HISTORICO_FILE, "w", encoding="utf-8") as f:
        json.dump(historico, f, indent=2, ensure_ascii=False)

def abrir_tela_historico():
    """Abre janela para visualizar histórico de comandos"""
    tema = TEMAS[tema_atual]
    
    tela_historico = tk.Toplevel(root)
    tela_historico.title("📜 Histórico de Comandos")
    tela_historico.geometry("900x700")
    tela_historico.configure(bg=tema["bg"])
    tela_historico.minsize(700, 550)
    
    tela_historico.transient(root)
    tela_historico.grab_set()
    
    main_frame = ttk.Frame(tela_historico, padding=20)
    main_frame.pack(fill="both", expand=True)
    
    titulo = ttk.Label(main_frame, text="Histórico de Execuções", style="Title.TLabel")
    titulo.pack(pady=(0, 15))
    
    # Frame para Treeview
    tree_frame = ttk.Frame(main_frame)
    tree_frame.pack(fill="both", expand=True)
    
    # Treeview para histórico
    colunas = ("timestamp", "status", "registros", "comando")
    tree_hist = ttk.Treeview(tree_frame, columns=colunas, show="headings", height=15)
    
    tree_hist.heading("timestamp", text="Data/Hora", anchor="w")
    tree_hist.heading("status", text="Status", anchor="center")
    tree_hist.heading("registros", text="Registros", anchor="center")
    tree_hist.heading("comando", text="Comando", anchor="w")
    
    tree_hist.column("timestamp", width=150, stretch=False)
    tree_hist.column("status", width=70, stretch=False, anchor="center")
    tree_hist.column("registros", width=80, stretch=False, anchor="center")
    tree_hist.column("comando", width=500, stretch=True)
    
    # Scrollbars
    scroll_v = ttk.Scrollbar(tree_frame, orient="vertical", command=tree_hist.yview)
    scroll_h = ttk.Scrollbar(main_frame, orient="horizontal", command=tree_hist.xview)
    tree_hist.configure(yscrollcommand=scroll_v.set, xscrollcommand=scroll_h.set)
    
    tree_hist.pack(side="left", fill="both", expand=True)
    scroll_v.pack(side="right", fill="y")
    scroll_h.pack(fill="x")
    
    # Carregar histórico
    if os.path.exists(HISTORICO_FILE):
        try:
            with open(HISTORICO_FILE, "r", encoding="utf-8") as f:
                historico = json.load(f)
                for item in historico:
                    status = "✅" if item.get("sucesso", False) else "❌"
                    registros = str(item.get("registros", "-"))
                    comando_resumo = item.get("comando", "")[:100].replace("\n", " ")
                    if len(item.get("comando", "")) > 100:
                        comando_resumo += "..."
                    tree_hist.insert("", "end", values=(
                        item.get("timestamp", ""),
                        status,
                        registros,
                        comando_resumo
                    ))
        except Exception as e:
            messagebox.showerror("❌ Erro", f"Erro ao carregar histórico:\n{str(e)}", parent=tela_historico)
    
    def usar_comando():
        """Copia o comando selecionado para o editor"""
        selected = tree_hist.selection()
        if not selected:
            messagebox.showwarning("⚠️ Atenção", "Selecione um comando!", parent=tela_historico)
            return
        
        # Buscar comando completo do histórico
        item_index = tree_hist.index(selected[0])
        if os.path.exists(HISTORICO_FILE):
            with open(HISTORICO_FILE, "r", encoding="utf-8") as f:
                historico = json.load(f)
                if item_index < len(historico):
                    comando = historico[item_index].get("comando", "")
                    caixa_texto_comando.delete("1.0", tk.END)
                    caixa_texto_comando.insert("1.0", comando)
                    tela_historico.destroy()
    
    def limpar_historico():
        """Limpa todo o histórico"""
        if messagebox.askyesno("🗑️ Confirmar", "Limpar todo o histórico?", parent=tela_historico):
            if os.path.exists(HISTORICO_FILE):
                os.remove(HISTORICO_FILE)
            for item in tree_hist.get_children():
                tree_hist.delete(item)
            messagebox.showinfo("✅ Sucesso", "Histórico limpo!", parent=tela_historico)
    
    # Frame de botões
    btn_frame = ttk.Frame(main_frame)
    btn_frame.pack(fill="x", pady=(15, 0))
    
    btn_usar = ttk.Button(btn_frame, text="📋 Usar Comando", style="Accent.TButton", command=usar_comando)
    btn_usar.pack(side="left")
    
    btn_limpar = ttk.Button(btn_frame, text="🗑️ Limpar Histórico", command=limpar_historico)
    btn_limpar.pack(side="right")
    
    # Double-click para usar comando
    tree_hist.bind("<Double-1>", lambda e: usar_comando())

def executar_comando():
    comando = caixa_texto_comando.get("1.0", tk.END).strip()

    if not comando:
        messagebox.showwarning("⚠️ Atenção", "Nenhum comando para executar!")
        return

    dados_conexao = obter_conexao_ativa()
    if not dados_conexao:
        messagebox.showerror("❌ Erro", "Configure uma conexão primeiro!")
        return

    string_conexao = (
        f"DRIVER={{SQL Server}};"
        f"SERVER={dados_conexao.get('ip', '')};"
        f"DATABASE={dados_conexao.get('banco', '')};"
        f"UID={dados_conexao.get('usuario', '')};"
        f"PWD={dados_conexao.get('senha', '')};"
    )

    try:
        conexao = pyodbc.connect(string_conexao)
        cursor = conexao.cursor()
        cursor.execute(comando)

        if comando.strip().upper().startswith("SELECT"):
            resultados = cursor.fetchall()
            
            for row in treeview.get_children():
                treeview.delete(row)
            
            colunas = [column[0] for column in cursor.description]
            treeview["columns"] = colunas
            treeview.heading("#0", text="#", anchor="w")
            treeview.column("#0", width=50, stretch=False)
            
            for coluna in colunas:
                treeview.heading(coluna, text=coluna, anchor="w")
                treeview.column(coluna, width=120, stretch=True)
            
            for i, linha in enumerate(resultados):
                valores_formatados = [str(valor) if valor is not None else "NULL" for valor in linha]
                treeview.insert("", "end", text=str(i + 1), values=valores_formatados)
            
            # Ajustar largura das colunas
            fonte = font.Font(family="Segoe UI", size=9)
            for coluna in colunas:
                coluna_index = colunas.index(coluna)
                largura_cabecalho = fonte.measure(coluna) + 20
                largura_maxima = largura_cabecalho
                for linha in resultados[:50]:  # Limitar para performance
                    valor = str(linha[coluna_index]) if linha[coluna_index] is not None else "NULL"
                    largura_maxima = max(largura_maxima, fonte.measure(valor) + 20)
                treeview.column(coluna, width=min(largura_maxima, 300))
            
            # Atualizar label de status
            label_status.config(text=f"✅ {len(resultados)} registro(s) retornado(s)")
            salvar_historico(comando, True, "SELECT executado", len(resultados))
        else:
            conexao.commit()
            label_status.config(text="✅ Comando executado com sucesso!")
            messagebox.showinfo("✅ Sucesso", "Comando executado com sucesso!")
            salvar_historico(comando, True, "Comando DML executado")

        cursor.close()
        conexao.close()

    except Exception as e:
        label_status.config(text="❌ Erro na execução")
        messagebox.showerror("❌ Erro", f"Falha ao executar:\n{str(e)}")
        salvar_historico(comando, False, str(e)[:200])

def exportar_csv():
    """Exporta os resultados da Treeview para CSV"""
    # Verificar se há dados
    children = treeview.get_children()
    if not children:
        messagebox.showwarning("⚠️ Atenção", "Não há dados para exportar!")
        return
    
    # Solicitar local para salvar
    arquivo = filedialog.asksaveasfilename(
        defaultextension=".csv",
        filetypes=[("Arquivo CSV", "*.csv"), ("Todos os arquivos", "*.*")],
        title="Salvar como CSV",
        initialname=f"exportacao_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    )
    
    if not arquivo:
        return
    
    try:
        colunas = treeview["columns"]
        
        with open(arquivo, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f, delimiter=";")
            
            # Escrever cabeçalho
            writer.writerow(colunas)
            
            # Escrever dados
            for item in children:
                valores = treeview.item(item, "values")
                writer.writerow(valores)
        
        label_status.config(text=f"✅ Exportado: {os.path.basename(arquivo)}")
        messagebox.showinfo("✅ Sucesso", f"Dados exportados para:\n{arquivo}")
        
    except Exception as e:
        messagebox.showerror("❌ Erro", f"Falha na exportação:\n{str(e)}")

def exportar_excel():
    """Exporta os resultados da Treeview para Excel"""
    # Verificar se há dados
    children = treeview.get_children()
    if not children:
        messagebox.showwarning("⚠️ Atenção", "Não há dados para exportar!")
        return
    
    # Tentar importar openpyxl
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        messagebox.showerror(
            "❌ Erro", 
            "Biblioteca 'openpyxl' não instalada!\n\n"
            "Execute no terminal:\npip install openpyxl"
        )
        return
    
    # Solicitar local para salvar
    arquivo = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Arquivo Excel", "*.xlsx"), ("Todos os arquivos", "*.*")],
        title="Salvar como Excel",
        initialname=f"exportacao_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )
    
    if not arquivo:
        return
    
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Dados"
        
        colunas = treeview["columns"]
        
        # Estilo para cabeçalho
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
        # Escrever cabeçalho
        for col_idx, coluna in enumerate(colunas, 1):
            cell = ws.cell(row=1, column=col_idx, value=coluna)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Escrever dados
        for row_idx, item in enumerate(children, 2):
            valores = treeview.item(item, "values")
            for col_idx, valor in enumerate(valores, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=valor)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")
        
        # Ajustar largura das colunas
        for col_idx, coluna in enumerate(colunas, 1):
            max_length = len(str(coluna))
            for item in children:
                valores = treeview.item(item, "values")
                if col_idx <= len(valores):
                    max_length = max(max_length, len(str(valores[col_idx - 1])))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 2, 50)
        
        wb.save(arquivo)
        
        label_status.config(text=f"✅ Exportado: {os.path.basename(arquivo)}")
        messagebox.showinfo("✅ Sucesso", f"Dados exportados para:\n{arquivo}")
        
    except Exception as e:
        messagebox.showerror("❌ Erro", f"Falha na exportação:\n{str(e)}")

def limpar_tela():
    """Limpa a área de comando e resultados"""
    caixa_texto_comando.delete("1.0", tk.END)
    for row in treeview.get_children():
        treeview.delete(row)
    combo_comandos.set("")
    label_status.config(text="")

# ============================================
# INTERFACE PRINCIPAL
# ============================================

# Carregar preferência de tema antes de criar a janela
carregar_preferencia_tema()

# Criar janela principal
root = tk.Tk()
root.title("FAI-SQL")

# Dimensões e posicionamento
largura = 1000
altura = 750

largura_tela = root.winfo_screenwidth()
altura_tela = root.winfo_screenheight()

pos_x = (largura_tela // 2) - (largura // 2)
pos_y = (altura_tela // 2) - (altura // 2)

root.geometry(f"{largura}x{altura}+{pos_x}+{pos_y}")
root.minsize(800, 600)
root.resizable(True, True)

# Barra de menus
menubar = tk.Menu(root, borderwidth=0)

menu_arquivo = tk.Menu(menubar, tearoff=0)
menu_arquivo.add_command(label="🗑️ Limpar Tela", command=limpar_tela)
menu_arquivo.add_separator()
menu_arquivo.add_command(label="🚪 Sair", command=sair)
menubar.add_cascade(label="Arquivo", menu=menu_arquivo)

menu_sistema = tk.Menu(menubar, tearoff=0)
menu_sistema.add_command(label="⚙️ Conexão", command=abrir_tela_conexao)
menu_sistema.add_command(label="➕ Adicionar Comando", command=abrir_tela_adicionar_comando)
menu_sistema.add_command(label="📜 Histórico", command=abrir_tela_historico)
menu_sistema.add_separator()
texto_tema = "🌙 Tema Escuro" if tema_atual == "claro" else "☀️ Tema Claro"
menu_sistema.add_command(label=texto_tema, command=alternar_tema)
menubar.add_cascade(label="Sistema", menu=menu_sistema)

root.config(menu=menubar)

# Frame principal com padding
main_frame = ttk.Frame(root, padding=20)
main_frame.pack(fill="both", expand=True)

# Label de conexão ativa
label_conexao_ativa = ttk.Label(main_frame, text="⚠️ Nenhuma conexão configurada", 
                                 font=("Segoe UI", 9))
label_conexao_ativa.pack(anchor="e", pady=(0, 5))

# Inicializar conexões e atualizar label
carregar_conexoes()  # Migra formato antigo se necessário
root.after(100, atualizar_label_conexao)  # Atualizar após carregar

# Frame superior - Seleção de comandos
frame_comandos = ttk.LabelFrame(main_frame, text="Comandos Salvos", padding=15)
frame_comandos.pack(fill="x", pady=(0, 15))

frame_combo = ttk.Frame(frame_comandos)
frame_combo.pack(fill="x")

combo_comandos = ttk.Combobox(frame_combo, state="readonly", font=("Segoe UI", 10))
combo_comandos.pack(side="left", fill="x", expand=True, padx=(0, 10))
combo_comandos.bind("<<ComboboxSelected>>", ao_selecionar_comando)

# Botões de gerenciamento de comandos
btn_editar = ttk.Button(frame_combo, text="✏️ Editar", command=abrir_tela_editar_comando)
btn_editar.pack(side="left", padx=(0, 5))

btn_excluir = ttk.Button(frame_combo, text="🗑️ Excluir", command=excluir_comando)
btn_excluir.pack(side="left")

# Frame do editor SQL
frame_editor = ttk.LabelFrame(main_frame, text="Editor SQL", padding=15)
frame_editor.pack(fill="x", pady=(0, 15))

tema = TEMAS[tema_atual]
caixa_texto_comando = tk.Text(frame_editor, height=8, wrap=tk.WORD,
                               font=("Consolas", 11),
                               bg=tema["entry_bg"], fg=tema["fg"],
                               insertbackground=tema["fg"], relief="flat",
                               highlightthickness=1, highlightbackground=tema["border"],
                               highlightcolor=tema["accent"],
                               undo=True)  # Habilitar undo
caixa_texto_comando.pack(fill="x")

# Configurar syntax highlighting
configurar_tags_syntax(caixa_texto_comando)
caixa_texto_comando.bind("<KeyRelease>", aplicar_syntax_highlighting)

# Configurar Undo/Redo (Ctrl+Z e Ctrl+Y)
def executar_undo(event=None):
    try:
        caixa_texto_comando.edit_undo()
        aplicar_syntax_highlighting()
    except tk.TclError:
        pass  # Nada para desfazer
    return "break"

def executar_redo(event=None):
    try:
        caixa_texto_comando.edit_redo()
        aplicar_syntax_highlighting()
    except tk.TclError:
        pass  # Nada para refazer
    return "break"

caixa_texto_comando.bind("<Control-z>", executar_undo)
caixa_texto_comando.bind("<Control-Z>", executar_undo)
caixa_texto_comando.bind("<Control-y>", executar_redo)
caixa_texto_comando.bind("<Control-Y>", executar_redo)

# Frame de botões de ação
frame_acoes = ttk.Frame(main_frame)
frame_acoes.pack(fill="x", pady=(0, 15))

btn_executar = ttk.Button(frame_acoes, text="▶️ Executar", 
                          style="Accent.TButton", command=executar_comando)
btn_executar.pack(side="left")

btn_limpar = ttk.Button(frame_acoes, text="🗑️ Limpar", command=limpar_tela)
btn_limpar.pack(side="left", padx=(10, 0))

# Separador visual
ttk.Separator(frame_acoes, orient="vertical").pack(side="left", padx=15, fill="y")

# Botões de exportação
btn_csv = ttk.Button(frame_acoes, text="📄 CSV", command=exportar_csv)
btn_csv.pack(side="left")

btn_excel = ttk.Button(frame_acoes, text="📊 Excel", command=exportar_excel)
btn_excel.pack(side="left", padx=(5, 0))

label_status = ttk.Label(frame_acoes, text="")
label_status.pack(side="right")

# Frame de resultados
frame_resultados = ttk.LabelFrame(main_frame, text="Resultados", padding=15)
frame_resultados.pack(fill="both", expand=True)

# Frame interno para Treeview e scrollbars
tree_frame = ttk.Frame(frame_resultados)
tree_frame.pack(fill="both", expand=True)

treeview = ttk.Treeview(tree_frame, show="headings")
treeview.pack(side="left", fill="both", expand=True)

scrollbar_vertical = ttk.Scrollbar(tree_frame, orient="vertical", command=treeview.yview)
scrollbar_vertical.pack(side="right", fill="y")
treeview.configure(yscrollcommand=scrollbar_vertical.set)

scrollbar_horizontal = ttk.Scrollbar(frame_resultados, orient="horizontal", command=treeview.xview)
scrollbar_horizontal.pack(side="bottom", fill="x")
treeview.configure(xscrollcommand=scrollbar_horizontal.set)

# Carregar comandos e aplicar tema
carregar_comandos_no_combobox()
aplicar_tema()

# Iniciar aplicação
root.mainloop()